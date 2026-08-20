#!/usr/bin/env python3
"""Gera o site estático da aba MOSTRUARIO da planilha de preços."""

from __future__ import annotations

import argparse
import html
import os
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_NAME = "MOSTRUARIO"
MAX_COLUMNS = 12
MAX_SCAN_ROWS = 10_000
EMPTY_ROWS_TO_STOP = 25
PRICE_COLUMNS = {8, 9, 10, 11}  # I:L, índices 1-based


def recalculate_workbook(source: Path, workdir: Path) -> Path:
    """Recalcula o XLSX com LibreOffice e retorna o arquivo recalculado."""
    if shutil.which("soffice") is None:
        raise RuntimeError(
            "LibreOffice (soffice) é obrigatório para resolver as fórmulas da planilha."
        )

    recalculated_dir = workdir / "recalculated"
    recalculated_dir.mkdir()
    copied_source = workdir / source.name
    shutil.copy2(source, copied_source)

    env = os.environ.copy()
    env["HOME"] = str(workdir / "home")
    Path(env["HOME"]).mkdir()

    command = [
        "soffice",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(recalculated_dir),
        str(copied_source),
    ]
    result = subprocess.run(command, capture_output=True, text=True, env=env)
    recalculated = recalculated_dir / source.name
    if result.returncode != 0 or not recalculated.exists():
        raise RuntimeError(
            "Falha ao recalcular a planilha com LibreOffice.\n"
            f"stdout: {result.stdout}\n"
            f"stderr: {result.stderr}"
        )
    return recalculated


def read_mostruario(workbook_path: Path) -> tuple[list[str], list[list[Any]]]:
    workbook = load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"A aba obrigatória {SHEET_NAME!r} não existe na planilha.")

    worksheet = workbook[SHEET_NAME]
    headers = [worksheet.cell(1, column).value for column in range(1, MAX_COLUMNS + 1)]
    headers = [str(value or "") for value in headers]

    rows: list[list[Any]] = []
    empty_streak = 0
    scan_limit = min(worksheet.max_row, MAX_SCAN_ROWS)
    for row_number in range(2, scan_limit + 1):
        values = [worksheet.cell(row_number, column).value for column in range(1, MAX_COLUMNS + 1)]
        if not any(value not in (None, "") for value in values):
            empty_streak += 1
            if rows and empty_streak >= EMPTY_ROWS_TO_STOP:
                break
            continue
        empty_streak = 0
        rows.append(values)

    if not rows:
        raise ValueError(f"A aba {SHEET_NAME!r} não contém registros de produtos.")
    if len(rows) >= scan_limit and scan_limit == MAX_SCAN_ROWS:
        raise ValueError(
            f"A leitura atingiu o limite de {MAX_SCAN_ROWS} linhas; revise o contrato da aba."
        )
    return headers, rows


def format_value(value: Any, column_number: int) -> str:
    if value is None:
        return ""
    if column_number in PRICE_COLUMNS and isinstance(value, (int, float)):
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return str(value)


def render_html(headers: list[str], rows: list[list[Any]], source_name: str) -> str:
    generated_at = datetime.now(timezone.utc).astimezone().strftime("%d/%m/%Y %H:%M:%S %Z")
    table_rows = []
    card_rows = []

    def detail_block(header: str, value: Any, position: int) -> str:
        return (
            '<div class="field">'
            f'<span class="field-label">{html.escape(header)}</span>'
            f'<span class="field-value{ " price" if position in PRICE_COLUMNS else "" }">'
            f'{html.escape(format_value(value, position))}</span>'
            '</div>'
        )

    for row in rows:
        search_text = " ".join(str(value or "") for value in row).lower()

        cells = []
        for position, value in enumerate(row, start=1):
            css_class = " class=\"price\"" if position in PRICE_COLUMNS else ""
            cells.append(f"<td{css_class}>{html.escape(format_value(value, position))}</td>")
        table_rows.append(
            f'<tr class="product-item" data-search="{html.escape(search_text, quote=True)}">{"".join(cells)}</tr>'
        )

        sku = format_value(row[1], 2)
        model = format_value(row[6], 7)
        brand = format_value(row[7], 8)
        avista = format_value(row[8], 9)
        x3 = format_value(row[9], 10)
        x6 = format_value(row[10], 11)
        x10 = format_value(row[11], 12)

        summary_cols = []
        for label, value in ((headers[0], row[0]), (headers[1], row[1]), (headers[6], row[6]), (headers[7], row[7])):
            summary_cols.append(
                '<div class="summary-chip">'
                f'<span>{html.escape(label)}</span>'
                f'<strong>{html.escape(format_value(value, 1))}</strong>'
                '</div>'
            )

        details = []
        for position, (header, value) in enumerate(zip(headers, row), start=1):
            details.append(detail_block(header, value, position))

        card_rows.append(
            f'''
<article class="card product-item" data-search="{html.escape(search_text, quote=True)}">
  <div class="card-head">
    <div>
      <div class="card-label">CS-LOJA</div>
      <div class="card-title">{html.escape(format_value(row[0], 1))}</div>
    </div>
    <div>
      <div class="card-label">Marca</div>
      <div class="card-title">{html.escape(brand)}</div>
    </div>
  </div>
  <div class="card-subtitle">{html.escape(sku)} · {html.escape(model)}</div>
  <div class="summary-grid">{''.join(summary_cols)}</div>
  <div class="price-grid">
    <div class="price-item"><span>À vista</span><strong>{html.escape(avista)}</strong></div>
    <div class="price-item"><span>3x</span><strong>{html.escape(x3)}</strong></div>
    <div class="price-item"><span>6x</span><strong>{html.escape(x6)}</strong></div>
    <div class="price-item"><span>10x</span><strong>{html.escape(x10)}</strong></div>
  </div>
  <details>
    <summary>Ver todos os campos</summary>
    <div class="details-grid">{''.join(details)}</div>
  </details>
</article>
'''.strip()
        )

    table_header = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    table_body = "\n".join(table_rows)
    cards_body = "\n".join(card_rows)
    escaped_source = html.escape(source_name)

    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Mostruário de preços</title>
  <style>
    :root {{ color-scheme: light; font-family: Inter, system-ui, -apple-system, sans-serif; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: #f4f6f8; color: #18212b; }}
    header {{ padding: 28px clamp(16px, 4vw, 56px) 20px; background: #102a43; color: white; }}
    h1 {{ margin: 0 0 8px; font-size: clamp(1.5rem, 3vw, 2.2rem); }}
    .meta {{ margin: 0; opacity: .8; font-size: .9rem; }}
    main {{ padding: 24px clamp(16px, 4vw, 56px) 48px; }}
    .toolbar {{ display: flex; gap: 12px; align-items: center; margin-bottom: 16px; flex-wrap: wrap; }}
    input {{ width: min(620px, 100%); padding: 12px 14px; border: 1px solid #bcccdc; border-radius: 10px; font-size: 1rem; }}
    #count {{ color: #52606d; font-size: .92rem; }}
    .table-wrap {{ overflow: auto; background: white; border: 1px solid #d9e2ec; border-radius: 10px; box-shadow: 0 2px 8px #102a4312; }}
    table {{ border-collapse: collapse; width: 100%; min-width: 980px; }}
    th, td {{ padding: 11px 12px; border-bottom: 1px solid #e6edf3; text-align: left; white-space: nowrap; }}
    th {{ position: sticky; top: 0; background: #eaf2f8; color: #243b53; font-size: .82rem; text-transform: uppercase; letter-spacing: .03em; }}
    td.price {{ text-align: right; font-variant-numeric: tabular-nums; }}
    tbody tr:hover {{ background: #f0f7ff; }}
    tbody tr:last-child td {{ border-bottom: 0; }}
    .cards {{ display: none; gap: 14px; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); }}
    .card {{ background: white; border: 1px solid #d9e2ec; border-radius: 14px; box-shadow: 0 2px 8px #102a430f; padding: 14px; }}
    .card-head {{ display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; margin-bottom: 8px; }}
    .card-label {{ font-size: .72rem; text-transform: uppercase; letter-spacing: .04em; color: #627d98; margin-bottom: 4px; }}
    .card-title {{ font-size: 1rem; font-weight: 700; line-height: 1.25; word-break: break-word; }}
    .card-subtitle {{ color: #52606d; font-size: .92rem; margin-bottom: 12px; word-break: break-word; }}
    .summary-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 8px; margin-bottom: 12px; }}
    .summary-chip, .price-item {{ background: #f8fafc; border: 1px solid #e6edf3; border-radius: 10px; padding: 10px 11px; }}
    .summary-chip span, .price-item span {{ display: block; font-size: .72rem; text-transform: uppercase; letter-spacing: .04em; color: #627d98; margin-bottom: 4px; }}
    .summary-chip strong, .price-item strong {{ display: block; font-size: .98rem; word-break: break-word; }}
    .price-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 8px; margin-bottom: 12px; }}
    .price-item strong {{ text-align: right; font-variant-numeric: tabular-nums; }}
    details {{ border-top: 1px solid #e6edf3; padding-top: 10px; }}
    summary {{ cursor: pointer; color: #334e68; font-weight: 600; }}
    .details-grid {{ display: grid; grid-template-columns: 1fr; gap: 8px; margin-top: 10px; }}
    .field {{ display: flex; justify-content: space-between; gap: 14px; background: #f8fafc; border: 1px solid #e6edf3; border-radius: 10px; padding: 10px 11px; }}
    .field-label {{ color: #627d98; font-size: .76rem; text-transform: uppercase; letter-spacing: .04em; flex: 1 1 auto; }}
    .field-value {{ font-weight: 600; text-align: right; flex: 0 0 auto; }}
    .field-value.price {{ font-variant-numeric: tabular-nums; }}
    .empty {{ padding: 28px; color: #52606d; }}
    @media (max-width: 980px) {{
      .table-wrap {{ display: none; }}
      .cards {{ display: grid; }}
      main {{ padding-bottom: 28px; }}
    }}
    @media (max-width: 640px) {{
      header {{ padding-bottom: 16px; }}
      main {{ padding-top: 18px; padding-left: 12px; padding-right: 12px; }}
      .summary-grid, .price-grid {{ grid-template-columns: 1fr; }}
      input {{ width: 100%; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Mostruário de preços</h1>
    <p class="meta">Fonte: {escaped_source} · Gerado em {html.escape(generated_at)}</p>
  </header>
  <main>
    <div class="toolbar">
      <input id="search" type="search" placeholder="Buscar por código, SKU, modelo, marca..." autocomplete="off">
      <span id="count"></span>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>{table_header}</tr></thead>
        <tbody id="products-table">{table_body}</tbody>
      </table>
    </div>
    <div id="cards" class="cards">{cards_body}</div>
    <div id="empty" class="empty" hidden>Nenhum produto encontrado.</div>
  </main>
  <script>
    const search = document.querySelector('#search');
    const tableRows = [...document.querySelectorAll('#products-table .product-item')];
    const cards = [...document.querySelectorAll('#cards .product-item')];
    const count = document.querySelector('#count');
    const empty = document.querySelector('#empty');
    const mobileQuery = window.matchMedia('(max-width: 980px)');

    function currentItems() {{
      return mobileQuery.matches ? cards : tableRows;
    }}

    function update() {{
      const query = search.value.trim().toLowerCase();
      const items = currentItems();
      let visible = 0;

      for (const item of tableRows) {{
        const show = !query || item.dataset.search.includes(query);
        item.hidden = mobileQuery.matches || !show;
        if (show && !mobileQuery.matches) visible++;
      }}
      for (const item of cards) {{
        const show = !query || item.dataset.search.includes(query);
        item.hidden = !mobileQuery.matches || !show;
        if (show && mobileQuery.matches) visible++;
      }}

      count.textContent = `${{visible}} de ${{items.length}} produtos`;
      empty.hidden = visible !== 0;
    }}

    search.addEventListener('input', update);
    mobileQuery.addEventListener('change', update);
    update();
  </script>
</body>
</html>
"""


def build(input_path: Path, output_dir: Path) -> int:
    input_path = input_path.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="planilia-web-") as temporary:
        recalculated = recalculate_workbook(input_path, Path(temporary))
        headers, rows = read_mostruario(recalculated)

    output_path = output_dir / "index.html"
    output_path.write_text(render_html(headers, rows, input_path.name), encoding="utf-8")
    print(f"Gerado: {output_path}")
    print(f"Aba: {SHEET_NAME}")
    print(f"Produtos: {len(rows)}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Caminho do XLSX fonte")
    parser.add_argument("--output", type=Path, required=True, help="Diretório do site gerado")
    args = parser.parse_args()
    return build(args.input, args.output)


if __name__ == "__main__":
    raise SystemExit(main())
